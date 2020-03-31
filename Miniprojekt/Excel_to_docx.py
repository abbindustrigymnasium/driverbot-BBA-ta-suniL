import openpyxl

wb = openpyxl.load_workbook('mello.xlsx')
sheet = wb.active

artister = [
                {
                    "Namn":sheet.cell(row=i, column=2).value,
                    "Sång":sheet.cell(row=i, column=3).value,
                    "Poäng":sheet.cell(row=i, column=6).value,
                    "Röst":sheet.cell(row=i, column=5).value,
                }    
                for i in range (2, sheet.max_row)
            ]

#print(artister)

vidare = [
                {
                    "Namn":sheet.cell(row=i, column=2).value,
                    "Sång":sheet.cell(row=i, column=3).value,
                    "Poäng":sheet.cell(row=i, column=6).value,
                    "Röst":sheet.cell(row=i, column=5).value,
                }    
                for i in range (2, sheet.max_row)
            
        if not sheet.cell(row=i, column=8).value.startswith("Utsl")
]

vidare = sorted(vidare, key=lambda k: k['Poäng'])
vidare = sorted(vidare, key=lambda k: k['Poäng'], reverse=True)
vidare = vidare[0:10]

#print(vidare)



[print(antal["Namn"]," är med i år!") for antal in artister]
[print(antal["Namn"]," var en av de",len(vidare)," som gick vidare med låten",antal["Sång"]) for antal in vidare]