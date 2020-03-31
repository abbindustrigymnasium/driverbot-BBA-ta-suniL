import openpyxl

wb = openpyxl.load_workbook('Farmen.xlsx')
sheet = wb.active

print(wb.get_sheet_names())

Deltagare = wb.get_sheet_by_name('Deltagare')
Veckosammanfattning = wb.get_sheet_by_name('Veckosammanfattning')
Tittarsiffror = wb.get_sheet_by_name('Tittarsiffror')


sheet = wb["Deltagare"]
sheet2 = wb["Veckosammanfattning"]
sheet3 = wb["Tittarsiffror"]



deltagare = [
    {
        "Namn":sheet.cell(row=i, column=1).value,
        "Ålder":sheet.cell(row=i, column=2).value,
        "Ort":sheet.cell(row=i, column=3).value,
        "Arbete":sheet.cell(row=i, column=4).value,
    }
    for i in range(2, sheet.max_row)
]

#[print(delt["Namn"], " är ", delt["Ålder"], " år gammal, kommer från ", delt["Ort"], " och jobbar som ", delt["Arbete"], ".") for delt in deltagare]
print('\n')






wkamper = [
    {
        sheet2.cell(row=i, column=5).value:0 for i in range(2, sheet2.max_row)
    }   
]


kamper = [sheet2.cell(row=antal, column=5).value for antal in range(2, sheet2.max_row)]


for i in range(2, sheet2.max_row):
    wkamper[0][sheet2.cell(row=i, column=5).value] += 1

print(max(wkamper[0], key=wkamper[0].get))

print('\n')





